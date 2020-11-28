import os
import openpyxl as opxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import time

#--------------Abre Excel -----------
def abrirexcel():
    os.system('start excel.exe "%s\\ezsave.xlsx"' % (sys.path[0],))

#-----------Activador de hoja en base al nombre ---------
def acth(nm, wb):
    hojas = wb.sheetnames
    x = len(hojas)
    for i in range(0, x):
        if nm in hojas[i]:
            return i
    return 0

#-----------Nueva hoja con nombre para un nuevo usuario------------
def newh(nm, p, wb):
    wb.create_sheet(nm)
    wb.save(p)
    wb.close()
    w2 = opxl.load_workbook(p)
    return w2

#-------------Creacion de hoja con formato para un nuevo usuario ----------
def formath(wba, p, wb):
    wb.active = wba
    Hoja = wb.active
    formato = ['META ESTADO', 'FECHA', 'META', 'PRESUPUESTO INICIAL', 'GASTO FIJO', 'INGRESO V', 'GASTO V']
    estado = [0]
    Hoja.append(formato)
    Hoja.append(estado)
    Hoja['H1'] = 0
    Hoja['I1'] = 0
    wb.save(p)
    wb.close()
    w2 = opxl.load_workbook(p)
    return w2

#----------Listas de el Index-------------
def listanh(wb):
    wb.active = 0
    wba = wb.active
    z = []
    w = wba.max_row
    for i in range(w):
        z.append(wba.cell(row=i + 1, column=2).value)

    return z


def listah(wb):
    wb.active = 0
    wba = wb.active
    x = []
    u = wba.max_row
    for i in range(u):
        x.append(wba.cell(row=i + 1, column=3).value)

    return x


def listach(wb):
    wb.active = 0
    wba = wb.active
    y = []
    v = wba.max_row
    for i in range(v):
        y.append(wba.cell(row=i + 1, column=4).value)

    return y


# -------Bienvenida-----------
def Bienvenida():
    print("---------------------------------")
    print("-------Bienvenido a EZSave-------")
    print("---------------------------------")
    print()
    print("Presiona Enter para continuar.")
    enter = input()


# -------Funciones de menus---------
# Eleccion Menu principal
def ElecMenuP():
    print("------------Menú---------------")
    print()
    print("Iniciar sesión................1")
    print()
    print("Crear un usuario..............2")
    print()
    print("Salir.........................3")
    print()
    print("Elige una opción:")
    op = int(input())
    return op


# Menu crear usuario
def MenuCU():
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    wb.active = 0
    Hoja = wb.active
    print("------Menú Crear Usuario------")
    print()
    Nuevo_Nombre = input("Escribe tu nombre y apellido: ")
    Nuevo_Nombre_de_Usuario = input("Escribe tu nombre de usuario: ")
    Nueva_Contrasena = input("Escribe tu contraseña: ")
    u = Hoja.max_row
    x = listah(wb)
    for c in range(len(x)):
        while Nuevo_Nombre_de_Usuario in x:
            Nuevo_Nombre_de_Usuario = input("Ese nombre de usuario ya existe, escribe uno diferente: ")

        Hoja.cell(row=(u + 1), column=2).value = Nuevo_Nombre
        Hoja.cell(row=(u + 1), column=3).value = Nuevo_Nombre_de_Usuario
        Hoja.cell(row=(u + 1), column=4).value = Nueva_Contrasena

    x = listah(wb)

    wb.save(path)
    wb.close()
    w2 = opxl.load_workbook(path)

    newh(Nuevo_Nombre_de_Usuario, path, wb)
    act = acth(Nuevo_Nombre_de_Usuario, wb)
    w2 = formath(act, path, wb)

    wb.save(path)
    wb.close()
    w2 = opxl.load_workbook(path)
    return w2


# -------Menu iniciar sesion------
def MenuIS():
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    wb.active = 0
    Hoja = wb.active
    u = Hoja.max_row

    # X es la lista de los nombres de usuario
    x = listah(wb)

    # y es la lista de las contrasenas
    y = listach(wb)

    intentos = 3
    print("------Menú iniciar Sesion------")
    print()
    Nombre_de_Usuario = input("Usuario: ")
    Contrasena = input("Contraseña: ")
    lp = 0
    while Nombre_de_Usuario not in x:
        lp = lp + 1
        Nombre_de_Usuario = input("Usuario no existe, introduce un nombre de usuario valido: ")
        if lp > 3:
            print("Excediste los intentos, cerraremos el programa.")
            exit()
    for i in range(len(x)):
        if Nombre_de_Usuario == x[i]:
            print('Usuario Correcto!')
            break

    if Contrasena != y[i]:

        while Contrasena != y[i]:
            print('Contraseña incorrecta, vuelve a intentar.')
            print(f'Te quedan {intentos} intentos')
            Contrasena = input("Introduce una contraseña valida: ")
            intentos = intentos - 1
            if intentos == 0:
                print('Intentaste entrar en una cuenta que no es tuya, se cerrará el programa.')
                exit()
    print('Contraseña correcta!')

    return Nombre_de_Usuario

#-------------Menu de usuario completo ---------
def MenuUs2(nm, wb, path, xx, xxx):
    tpl = Math(nm, wb, xx, xxx)
    # Formato
    sq = tpl[5]
    pr = ProgressB(sq)

    print("*" * 75)
    print("                                 EzSave \n\t                 Sistema de Control Financiero")
    print(" " * 26, time.strftime("%d/%m/%y"), "-", time.strftime("%I:%M:%S"))
    print(f"Tu meta es: {tpl[0]} ", end="")
    print(f"Tu porcentaje de avance es: {pr[0]}%[{pr[2]}{pr[3]}]%100")
    print(f"Tu ahorro hasta la fecha es de: {tpl[8]}")
    print("*" * 75)
    print()
    print("-" * 31, end="")
    print("Menú Usuario", end="")
    print("-" * 32)
    # Formato
    print()
    print("Nuevos Ingresos...........................................................1")
    print()
    print("Nuevos Egresos............................................................2")
    print()
    print("Acceso a Resumen..........................................................3")
    print()
    print('Entrar a base de datos ("Cuidado, NO TOCAR N A D A")......................4')
    print()
    print("Salir.....................................................................5")
    print("Elige una opción:")
    op = int(input())
    while op < 1 or op > 5:
        print("Esa no es una opción válida, por favor selecciona una opción valida.")
        op = int(input())
        print()
    if op == 1:  # Nuevos Ingresos
        w2 = NuevosI(nm)
        wb = w2
        xx = SuLiEg(nm, wb, path)
        xxx = SuLiIn(nm, wb, path)
        print()
        MenuUs2(nm, wb, path, xx, xxx)
    elif op == 2:  # Nuevos Egresos
        w2 = NuevosG(nm)
        wb = w2
        xx = SuLiEg(nm, wb, path)
        xxx = SuLiIn(nm, wb, path)
        print()
        MenuUs2(nm, wb, path, xx, xxx)
    elif op == 3:  # Acceso a Resumen
        Summary(tpl)
        print()
        MenuUs2(nm, wb, path, xx, xxx)

    elif op == 4:  # Entrar a la base de datos
        print()
        abrirexcel()
        MenuUs2(nm, wb, path, xx, xxx)

    elif op == 5:  # Salir
        print("Decidiste salir.")


#--------Estado de Meta --------
def Mstate(nm):
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    x = acth(nm, wb)
    wb.active = x
    h = wb.active
    st = h['A2'].value
    return st

#---------Recopilacion de datos para nuevos usuarios -------
def RecopData(wba):
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    wb.active = wba
    h = wb.active
    print()
    z = input("Meta: ")
    print()
    y = input("Presupuesto Inicial: ")
    print()
    w = input("Gasto Inicial: ")
    print()
    h['A2'] = 1  # Estado de la Meta
    h['C2'] = z  # Meta
    h['D2'] = y  # PresupuestoI
    h['E2'] = w  # GastoF

    wb.save(path)
    wb.close()
    w2 = opxl.load_workbook(path)

    return z, y, w, w2

#----------Inicio de Sesion Completo --------
def ISV():
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    us = MenuIS()
    st = Mstate(us)
    if st == 0:
        act = acth(us, wb)
        datos = RecopData(act)
        wb = datos[3]
        xx = SuLiEg(us, wb, path)
        xxx = SuLiIn(us, wb, path)
        MenuUs2(us, wb, path, xx, xxx)
    else:
        xx = SuLiEg(us, wb, path)
        xxx = SuLiIn(us, wb, path)
        MenuUs2(us, wb, path, xx, xxx)

    return wb

#---------Nuevos Ingresos ----------
def NuevosI(nm):
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    wba = acth(nm, wb)
    wb.active = wba
    h = wb.active
    col = h['F']
    lencol = len(col)
    x = []
    for i in col:
        x = x + [i.value]
        if None in x:
            x.remove(None)

    ni = int(input("Introduce tus nuevos ingresos, si ya no tienes nuevos ingresos escribe 0: "))
    for i in range(lencol):
        c = 0
        if h['H1'].value == 0:
            while ni != 0:
                h.cell(row=lencol + c, column=6).value = ni
                ni = int(input("Introduce tus nuevos ingresos, si ya no tienes nuevos ingresos escribe 0: "))
                c += 1
            h['H1'] = 1
        else:
            c = 1
            while ni != 0:
                h.cell(row=lencol + c, column=6).value = ni
                ni = int(input("Introduce tus nuevos ingresos, si ya no tienes nuevos ingresos escribe 0: "))
                c += 1

    wb.save(path)
    wb.close()
    w2 = opxl.load_workbook(path)

    return w2

#-------------Nuevos Egresos--------------
def NuevosG(nm):
    path = 'ezsave.xlsx'
    wb = opxl.load_workbook(path)
    wba = acth(nm, wb)
    wb.active = wba
    h = wb.active
    col = h['G']
    lencol = len(col)
    x = []
    k = -1
    for i in col:
        k = k + 1
        x = x + [i.value]
        if None in x:
            x.remove(None)

    ni = int(input("Introduce tus nuevos gastos, si ya no tienes nuevos gastos escribe 0: "))
    for i in range(lencol):
        c = 0
        if h['I1'].value == 0:
            while ni != 0:
                h.cell(row=len(x) + 1 + c, column=7).value = ni
                ni = int(input("Introduce tus nuevos gastos, si ya no tienes nuevos gastos escribe 0: "))
                c += 1
            h['I1'] = 1
        else:
            c = 1
            while ni != 0:
                h.cell(row=len(x) + c, column=7).value = ni
                ni = int(input("Introduce tus nuevos gastos, si ya no tienes nuevos gastos escribe 0: "))
                c += 1

    wb.save(path)
    wb.close()
    w2 = opxl.load_workbook(path)

    return w2

#--------------Sumatoria de Ingresos-----------
def SuLiIn(nm, wb, path):
    wba = acth(nm, wb)
    wb.active = wba
    h = wb.active
    col = h['F']
    xxx = []
    s = 0
    for i in col:
        xxx = xxx + [i.value]
        if None in xxx:
            xxx.remove(None)
    del xxx[0]
    for c in range(len(xxx)):
        s = s + xxx[c]
    return s

#----------------Sumatoria de Egresos------------
def SuLiEg(nm, wb, path):
    wba = acth(nm, wb)
    wb.active = wba
    h = wb.active
    col = h['G']
    xx = []
    k = 0
    for i in col:
        xx = xx + [i.value]
        if None in xx:
            xx.remove(None)
    del xx[0]
    for c in range(len(xx)):
        k = k + xx[c]
    return k

#-----------------Procesos matematicos ---------------
def Math(nm, wb, s, k):
    act = acth(nm, wb)
    wb.active = act
    h = wb.active
    m = h['C2'].value
    m = int(m)
    inf = h['D2'].value
    inf = int(inf)
    gf = h['E2'].value
    gf = int(gf)

    d = m / ((inf + k) - (gf + s))  # Cuantos meses se va a tardar en ahorrar
    x = m / ((inf - gf) / 30)  # Cuantos dias se va a tardar en ahorrar
    y = m / x  # Cuanto tiene que ahorrar al dia
    z = m / d  # Cuanto ahorrar al mes
    w = (100 * ((inf + k) - (gf + s))) / m  # Porcentaje que se tiene hasta el momento
    l = ((inf + k) - (gf + s))
    o = (m - l)/12
    om = (m-l)/30
    return m, d, x, y, z, w, k, s, l, o, om

#--------------Barra de progreso -----------------
def ProgressB(sq):
    pr = 0
    wp = int(sq)
    if 0 < sq <= 5:
        sq = 0
    if 5 < sq <= 10:
        sq = 5
        pr = 1
    if 10 < sq <= 15:
        sq = 10
        pr = 2
    if 15 < sq <= 20:
        sq = 15
        pr = 3
    if 20 < sq <= 25:
        sq = 20
        pr = 4
    if 25 < sq <= 30:
        sq = 25
        pr = 5
    if 30 < sq <= 35:
        sq = 30
        pr = 6
    if 35 < sq <= 40:
        sq = 35
        pr = 7
    if 40 < sq <= 45:
        sq = 40
        pr = 8
    if 45 < sq <= 50:
        sq = 45
        pr = 9
    if 50 < sq <= 55:
        sq = 50
        pr = 10
    if 55 < sq <= 60:
        sq = 55
        pr = 11
    if 60 < sq <= 65:
        sq = 60
        pr = 12
    if 65 < sq <= 70:
        sq = 65
        pr = 13
    if 70 < sq <= 75:
        sq = 70
        pr = 14
    if 75 < sq <= 80:
        sq = 75
        pr = 15
    if 80 < sq <= 85:
        sq = 80
        pr = 16
    if 85 < sq <= 90:
        sq = 85
        pr = 17
    if 90 < sq <= 95:
        sq = 90
        pr = 18
    if 95 < sq <= 100:
        sq = 95
        pr = 19
    if sq > 100:
        sq = 100
        pr = 20

    prr = 20 - pr
    a = ('▒' * prr)
    x = ('▓' * pr)
    return wp, sq, x, a

#------------Resumen de la informacion--------------
def Summary(tpl):
    print("-" * 75)
    print("-" * 34, end="")
    print("Resúmen", end="")
    print("-" * 34)
    print("-" * 75)
    print()
    print(f"Meta: {tpl[0]}")
    print()
    print(f"Tu ahorro a la fecha es de: {tpl[8]}")
    if tpl[1] > 1:
        print()
        print(f"Te tardarás {int(tpl[1])} meses en ahorrar.")
    else:
        print()
        print(f"Te tardarás {tpl[2]} días en ahorrar.")
    print()
    print(f"Tienes que ahorrar {int(tpl[9])} al mes, para completar tu meta en un año.")
    print()
    print(f"Tienes que ahorrar {int(tpl[10])} al día para llegar a tu meta en un mes.")
    print()
    print(f"Tus ingresos variables hasta la fecha han sido de: {tpl[6]}")
    print()
    print(f"Tus egresos variables hasta la fecha han sido de: {tpl[7]}")
    print('Pulsa la tecla enter para salir al menú de usuario.')
    x = input()